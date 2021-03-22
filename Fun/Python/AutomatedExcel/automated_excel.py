import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_work_book(filename, sheetname):
    work_book = xl.load_workbook(filename)
    sheet = work_book[sheetname]
    # code block for customized functionality
    # for row in range(2, sheet.max_row + 1):
    work_book.save(filename)


# open work book
work_book = xl.load_workbook("example.xlsx")
# open sheet inside work book
sheet = work_book["Sheet1"]
# select cell in sheet
# cell = sheet.cell(1, 1)
# print the value of the cell
# print(cell.value)
# max_row gets the number of last row
# print(sheet.max_row)

# range[inclusive, exclusive)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

work_book.save("transactions_v2.xlsx")